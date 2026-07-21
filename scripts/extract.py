#!/usr/bin/env python3
"""
Extract the TA case export into the app's data files.

Usage:
    python scripts/extract.py path/to/sn_customerservice_case.xlsx

What it does:
  1. Reads the ServiceNow "Case Report" export (first sheet).
  2. Writes app/src/data/cases.json   — one record per request (minified).
  3. Writes app/src/data/today.json   — the "as of" date as an Excel serial,
     derived as floor(latest Created / Opened / Updated timestamp).
  4. Rewrites the hard-coded "as of …" date labels in the UI so they match.

The column mapping and transforms below were validated to reproduce the
committed cases.json byte-for-byte from the source export (zero diff), so a
re-run on the same file is a no-op. Dates are stored as Excel serial day
numbers to match how the rest of the app reads them.

Requires: openpyxl  (pip install openpyxl)
"""
import sys
import os
import re
import json
import math
import datetime

# ---- source layout -------------------------------------------------------
# The export has a one-cell title row ("Case Report") followed by a real
# header row, then the data. These are the 23 columns we expect, in order.
EXPECTED_HEADER = (
    'Case Report', 'Request Type', 'Expected Start Date', 'Expected Completion Date',
    'Office/Division', 'Region', 'Requested For', 'Requested by', 'Short description',
    'Description', 'Objectives', 'Language', 'Modality',
    'Global Practice and Cross Sectoral Teams', 'Primary Programme Offer',
    'Assigned to', 'Implementation Status', 'Created', 'Opened', 'Updated',
    'Resolved', 'Closed', 'First Response Time',
)

# Column indices (0-based) used when building each record.
C_ID, C_TYPE, C_XS, C_XC, C_OFFICE, C_REGION, C_REQFOR = 0, 1, 2, 3, 4, 5, 6
C_SHORT, C_DESC, C_OBJ, C_MODALITY = 8, 9, 10, 12
C_PRACTICE, C_OFFER, C_LEAD, C_STATUS = 13, 14, 15, 16
C_CREATED, C_OPENED, C_UPDATED, C_RESOLVED, C_CLOSED = 17, 18, 19, 20, 21

EPOCH = datetime.datetime(1899, 12, 30)  # Excel serial-date origin


def serial(v):
    """Excel serial day number for a datetime; pass through numbers; None stays None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return (v - EPOCH).total_seconds() / 86400.0


def s(v):
    return '' if v is None else str(v)


def load_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()  # the export stores a bogus A1 dimension
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = tuple(rows[0])
    if header != EXPECTED_HEADER:
        raise SystemExit(
            'ERROR: unexpected export layout.\n'
            f'  expected {len(EXPECTED_HEADER)} columns starting {EXPECTED_HEADER[:4]}\n'
            f'  got      {len(header)} columns starting {header[:4]}\n'
            'The column mapping in this script would need updating.'
        )
    return rows[1:]


def build(r):
    """One export row -> one TACase record (see app/src/data/types.ts)."""
    return {
        'id': s(r[C_ID]), 'type': s(r[C_TYPE]),
        'region': s(r[C_REGION]), 'office': s(r[C_OFFICE]),
        'practice': s(r[C_PRACTICE]), 'offer': s(r[C_OFFER]), 'modality': s(r[C_MODALITY]),
        'status': s(r[C_STATUS]), 'lead': s(r[C_LEAD]),
        'reqFor': s(r[C_REQFOR]), 'desc': s(r[C_SHORT]),
        # long Description, falling back to the short description when blank
        'full': s(r[C_DESC]) if r[C_DESC] else s(r[C_SHORT]),
        'xs': serial(r[C_XS]), 'xc': serial(r[C_XC]),
        'cr': serial(r[C_CREATED]), 'op': serial(r[C_OPENED]), 'up': serial(r[C_UPDATED]),
        'rs': serial(r[C_RESOLVED]), 'cl': serial(r[C_CLOSED]),
        'hd': 1 if r[C_DESC] else 0, 'ho': 1 if r[C_OBJ] else 0,
    }


def fmt_date(serial_day):
    d = EPOCH + datetime.timedelta(days=serial_day)
    return d.strftime('%-d %b %Y'), d


def update_labels(app_dir, today_serial, n_rows):
    """Rewrite the hard-coded 'as of …' date strings so the UI matches the data."""
    today_str, today_dt = fmt_date(today_serial)
    win_dt = today_dt - datetime.timedelta(days=30)
    win_full = win_dt.strftime('%-d %b %Y')   # e.g. "17 Jun 2026"
    win_short = win_dt.strftime('%-d %b')      # e.g. "17 Jun"
    n_str = f'{n_rows:,}'

    edits = [
        ('src/components/Header.tsx',
         r'as of \d+ \w+ \d{4}', f'as of {today_str}'),
        ('src/components/PerformanceView.tsx',
         r'between \d+ \w+ and \d+ \w+ \d{4}', f'between {win_short} and {today_str}'),
        ('src/components/PerformanceView.tsx',
         r'today \(\d+ \w+ \d{4}\)', f'today ({today_str})'),
        ('src/lib/dashboard.ts',
         r'new since \d+ \w+ \d{4}', f'new since {win_full}'),
        ('src/data/cases.ts',
         r'\([\d,]+ rows, as of \d+ \w+ \d{4}\)', f'({n_str} rows, as of {today_str})'),
    ]
    for rel, pat, repl in edits:
        path = os.path.join(app_dir, rel)
        with open(path, encoding='utf-8') as f:
            text = f.read()
        new_text, n = re.subn(pat, repl, text)
        if n == 0:
            print(f'  WARN: no label match in {rel} for /{pat}/')
        else:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(new_text)
    return today_str


def main():
    if len(sys.argv) != 2:
        raise SystemExit('Usage: python scripts/extract.py <export.xlsx>')
    src = sys.argv[1]
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    app_dir = os.path.join(repo_root, 'app')
    data_dir = os.path.join(app_dir, 'src', 'data')

    rows = load_rows(src)
    records = [build(r) for r in rows if s(r[C_ID])]

    # "as of" date = floor of the latest activity timestamp in the export
    stamps = [serial(r[i]) for r in rows for i in (C_CREATED, C_OPENED, C_UPDATED)]
    today = math.floor(max(v for v in stamps if v is not None))

    with open(os.path.join(data_dir, 'cases.json'), 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, separators=(',', ':'))
    with open(os.path.join(data_dir, 'today.json'), 'w', encoding='utf-8') as f:
        json.dump(today, f)

    today_str = update_labels(app_dir, today, len(records))

    from collections import Counter
    dist = dict(Counter(c['status'] for c in records))
    print(f'Wrote {len(records):,} records  ·  as of {today_str} (serial {today})')
    print(f'Status distribution: {dist}')


if __name__ == '__main__':
    main()
